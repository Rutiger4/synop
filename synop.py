import requests
import json
import csv
import openpyxl
from openpyxl.styles import Font
import datetime
import math
#import win32com.client as win32

from urllib3.exceptions import InsecureRequestWarning
# Suppress only the single warning from urllib3 needed.
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

STATIONS = [
    {"Name" : "Раздольное", "WMID" : "33922", "SYNOPID" : "5600"},
    {"Name" : "Черноморское", "WMID" : "33924", "SYNOPID" : "5601"},
    {"Name" : "Евпатория", "WMID" : "33929", "SYNOPID" : "5602"},
    {"Name" : "Ишунь", "WMID" : "33933", "SYNOPID" : "5603"},
    {"Name" : "Джанкой", "WMID" : "33934", "SYNOPID" : "5604"},
    {"Name" : "Клепинино", "WMID" : "33939", "SYNOPID" : "5605"},
    {"Name" : "Почтовое", "WMID" : "33945", "SYNOPID" : "5606"},
    {"Name" : "Симферополь аэропорт", "WMID" : "33946", "SYNOPID" : "5607"},
    {"Name" : "Симферополь город", "WMID" : "33955", "SYNOPID" : "5608"},
    {"Name" : "Карадаг", "WMID" : "33957", "SYNOPID" : "15752"},
    {"Name" : "Ангарский перевал", "WMID" : "33958", "SYNOPID" : "5609"},   
    {"Name" : "Алушта", "WMID" : "33959", "SYNOPID" : "5610"},
    {"Name" : "Нижнегорский", "WMID" : "33962", "SYNOPID" : "5613"},
    {"Name" : "Белогорск", "WMID" : "33966", "SYNOPID" : "5614"},
    {"Name" : "Владиславовка", "WMID" : "33973", "SYNOPID" : "5615"},
    {"Name" : "Феодосия", "WMID" : "33976", "SYNOPID" : "5616"},
    {"Name" : "Мысовое", "WMID" : "33981", "SYNOPID" : "5617"},
    {"Name" : "Керчь", "WMID" : "33983", "SYNOPID" : "5618"},
    {"Name" : "Опасное", "WMID" : "33986", "SYNOPID" : "5619"},
    {"Name" : "Ялта", "WMID" : "33990", "SYNOPID" : "5620"}, 
    {"Name" : "Севастополь", "WMID" : "33991", "SYNOPID" : "5621"}, 
    {"Name" : "Херсонесский Маяк", "WMID" : "33994", "SYNOPID" : "5622"}, 
    {"Name" : "Никита", "WMID" : "33995", "SYNOPID" : "5623"},
    {"Name" : "Ай-Петри", "WMID" : "33998", "SYNOPID" : "5624"}
]
METEOPARAMS = [1, 2, 4, 8, 16, 32, 128, 256, 512, 1024]
NAN = math.nan
EMPTYPARAMS = [NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN]
#cityID = "5600"
sDateStr = "2022-12-31T21:00:00"
eDateStr = "2022-12-01T00:00:00"
gmtD = datetime.timedelta(hours=3)
sDate = datetime.datetime.fromisoformat(sDateStr)
eDate = datetime.datetime.fromisoformat(eDateStr)
delta3h = datetime.timedelta(hours=3)

def loadFromSynop(isoStamp):
    synop = requests.get("https://synop.ru/forecast/SynopApp/ArchiveData?id=" + cityID + "&dateTime=" + isoStamp, verify=False)
    meteo = json.loads(synop.text)
    return meteo['MeteoData']

def writeCSV(wSelection):
    with open(cityID + '_' + sDateStr[0:10] + '_' + eDateStr[0:10] + '.csv', 'w', newline = '') as f:
        writer = csv.writer(f, delimiter = ';')
        for k,v in wSelection.items():
            row = [k]
            for i in range(len(v)):
                if math.isnan(v[i]):
                    row.append('')
                else:    
                    if i == 0:  
                        row.append('{:+.1f}'.format(v[i]).replace('.',','))
                    elif i == 2 or i == 6:
                        row.append('{:.1f}'.format(v[i]).replace('.',','))
                    else:
                        row.append('{:.0f}'.format(v[i]))                        
            writer.writerow(row)  

def writeXLS(wSelection, st):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.column_dimensions['A'].width = 20
    sheet.cell(row = 1, column = 1).value = st["Name"]
    sheet.cell(row = 1, column = 2).value = st["WMID"]
    sheet.cell(row = 1, column = 3).value = st["SYNOPID"]                       
    rowNum = 2;
    for k,v in wSelection.items():
        sheet.cell(row = rowNum, column = 1).value = k
        for i in range(len(v)):
            if not math.isnan(v[i]):  
                if i == 0:  
                    nform = '+0.0;-0.0'
                    #val = '{:+.1f}'.format(v[i]).replace('.',',')
                elif i == 2 or i == 6:
                    nform = '0.0'
                    #val = '{:.1f}'.format(v[i]).replace('.',',')
                else:
                    nform = '0'
                    #val = '{:.0f}'.format(v[i])   
                sheet.cell(row = rowNum, column = i+2).number_format = nform                    
                sheet.cell(row = rowNum, column = i+2).value = v[i]
        rowNum = rowNum + 1
    fontBold = Font(bold = True)
    for i in range(8):
        sheet.cell(row = rowNum+1, column = i+1).font = fontBold
    sheet.cell(row = rowNum+1, column = 2).number_format = '+0.00;-0.00' 
    sheet.cell(row = rowNum+2, column = 2).number_format = '+0.0;-0.0'
    sheet.cell(row = rowNum+3, column = 2).number_format = '+0.0;-0.0'         
    for j in range(3):
        sheet.cell(row = rowNum+1, column = j+3).number_format = '0.00'
    sheet.cell(row = rowNum+1, column = 1).value = 'TOTAL'   
    sheet.cell(row = rowNum+2, column = 1).value = 'MIN'   
    sheet.cell(row = rowNum+3, column = 1).value = 'MAX' 
    sheet.cell(row = rowNum+4, column = 1).value = 'NUM'                     
    sheet.cell(row = rowNum+1, column = 2).value = '=AVERAGE(B2:B' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+1, column = 3).value = '=AVERAGE(C2:C' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+1, column = 4).value = '=AVERAGE(D2:D' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+1, column = 5).value = '=AVERAGE(E2:E' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+1, column = 6).value = '=MAX(F2:F' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+1, column = 8).value = '=SUM(H2:H' + str(rowNum-1) + ')/2'
    sheet.cell(row = rowNum+2, column = 2).value = '=MIN(B2:B' + str(rowNum-1) + ')'  
    sheet.cell(row = rowNum+3, column = 2).value = '=MAX(B2:B' + str(rowNum-1) + ')'
    sheet.cell(row = rowNum+4, column = 2).value = '=COUNTA(B2:B' + str(rowNum-1) + ')'       

    wb.save(st["WMID"] + '_' + sDateStr[0:10] + '_' + eDateStr[0:10] + '.xlsx')
    '''
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open('D:\\ДОКИБУКИ\\Исследования, статистика\\Климат\\Крым обзоры\\synop\\' + st["WMID"] + '_' + sDateStr[0:10] + '_' + eDateStr[0:10] + '.xlsx')
    workbook.Save()
    workbook.Close()
    excel.Quit()        '''

def writeSummaryXLS():
    tb = openpyxl.Workbook()
    tsheet = tb.active
    tsheet.append(["WMID", "Num", "T", "Tn", "Tx", "U", "P", "FF", "F10", "RRR"])  
    for st in STATIONS:
        fileXLS = st["WMID"] + '_' + sDateStr[0:10] + '_' + eDateStr[0:10] + '.xlsx'
        wb = openpyxl.load_workbook(fileXLS, data_only=True)  
        sheet = wb["Sheet"]
        MR = sheet.max_row - 3
        begList = [st["WMID"], sheet.cell(row = MR+3, column = 2).value, sheet.cell(row = MR, column = 2).value, sheet.cell(row = MR+1, column = 2).value, sheet.cell(row = MR+2, column = 2).value]
        for i in range(3, 7):
            begList.append(sheet.cell(row = MR, column = i).value)   
        begList.append(sheet.cell(row = MR, column = 8).value)    
        tsheet.append(begList)
    tb.save("crimea" + "_" + eDateStr[0:7] + '.xlsx')

for city in STATIONS:
    cityID = city["SYNOPID"]
    print(city["Name"])
    selection = {}    
    cDate = sDate
    #Создаём словрь-болванку со всеми часами наблюдений в заданном интервале 
    while cDate >= eDate:
        selection[cDate.strftime('%Y-%m-%dT%H:00:00')] = EMPTYPARAMS.copy()
        cDate = cDate - delta3h
    #print(selection)
    #Заполняем словарь, подгружая данные из SYNOP.ru
    weatherData = {}
    cDate = sDate - gmtD;
    while cDate > (eDate - gmtD):
        while len(weatherData) > 0:
            lOK, lOV = weatherData.popitem()
            wDataDate = datetime.datetime.fromisoformat(lOK[:-1])
            localDateStr = (wDataDate + gmtD).strftime('%Y-%m-%dT%H:00:00')
            if localDateStr in selection:
                cDate = wDataDate
                unitList = EMPTYPARAMS.copy()
                for u in lOV:
                    unitList[METEOPARAMS.index(u['MeteoParam'])] = u['Value']
                    selection[localDateStr] = unitList
        cDateStr = cDate.strftime('%Y-%m-%dT%H:00:00')
        #print (cDateStr)
        weatherData = loadFromSynop(cDateStr)
    
    #print (selection)                       
    #writeCSV(selection)
    writeXLS(selection, city)

#writeSummaryXLS()

#for tstamp, observ in weatherData.items():
#    print(tstamp, end = "|")
#    for units in observ:
#        print(str(units['Value']), end = "|")
#    print("\n")

#print (prevObserv(date + "T" + time))      