import requests
import json
import csv
import openpyxl
from openpyxl.styles import Font
import datetime
import math
#import win32com.client as win32

from urllib3.exceptions import InsecureRequestWarning
# Suppress only the single warning from urllib3 needed.
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

STATIONS = [
    {"Name" : "Раздольное", "WMID" : "33922", "SYNOPID" : "5600"},
    {"Name" : "Черноморское", "WMID" : "33924", "SYNOPID" : "5601"},
    {"Name" : "Евпатория", "WMID" : "33929", "SYNOPID" : "5602"},
    {"Name" : "Ишунь", "WMID" : "33933", "SYNOPID" : "5603"},
    {"Name" : "Джанкой", "WMID" : "33934", "SYNOPID" : "5604"},
    {"Name" : "Клепинино", "WMID" : "33939", "SYNOPID" : "5605"},
    {"Name" : "Почтовое", "WMID" : "33945", "SYNOPID" : "5606"},
    {"Name" : "Симферополь аэропорт", "WMID" : "33946", "SYNOPID" : "5607"},
    {"Name" : "Симферополь город", "WMID" : "33955", "SYNOPID" : "5608"},
    {"Name" : "Карадаг", "WMID" : "33957", "SYNOPID" : "15752"},
    {"Name" : "Ангарский перевал", "WMID" : "33958", "SYNOPID" : "5609"},   
    {"Name" : "Алушта", "WMID" : "33959", "SYNOPID" : "5610"},
    {"Name" : "Нижнегорский", "WMID" : "33962", "SYNOPID" : "5613"},
    {"Name" : "Белогорск", "WMID" : "33966", "SYNOPID" : "5614"},
    {"Name" : "Владиславовка", "WMID" : "33973", "SYNOPID" : "5615"},
    {"Name" : "Феодосия", "WMID" : "33976", "SYNOPID" : "5616"},
    {"Name" : "Мысовое", "WMID" : "33981", "SYNOPID" : "5617"},
    {"Name" : "Керчь", "WMID" : "33983", "SYNOPID" : "5618"},
    {"Name" : "Опасное", "WMID" : "33986", "SYNOPID" : "5619"},
    {"Name" : "Ялта", "WMID" : "33990", "SYNOPID" : "5620"}, 
    {"Name" : "Севастополь", "WMID" : "33991", "SYNOPID" : "5621"}, 
    {"Name" : "Херсонесский Маяк", "WMID" : "33994", "SYNOPID" : "5622"}, 
    {"Name" : "Никита", "WMID" : "33995", "SYNOPID" : "5623"},
    {"Name" : "Ай-Петри", "WMID" : "33998", "SYNOPID" : "5624"}
]
METEOPARAMS = [1, 2, 4, 8, 16, 32, 128, 256, 512, 1024]
NAN = math.nan
EMPTYPARAMS = [NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN, NAN]
#cityID = "5600"
sDateStr = "2022-10-31T21:00:00"
eDateStr = "2022-10-01T00:00:00"
gmtD = datetime.timedelta(hours=3)

sDate = datetime.datetime.fromisoformat(sDateStr)
eDate = datetime.datetime.fromisoformat(eDateStr)
delta3h = datetime.timedelta(hours=3)

def loadFromSynop(isoStamp):
    synop = requests.get("https://synop.ru/forecast/SynopApp/ArchiveData?id=" + cityID + "&dateTime=" + isoStamp, verify=False)
    meteo = json.loads(synop.text)
    return meteo['MeteoData']

def writeSummaryXLS():
    tb = openpyxl.Workbook()
    tsheet = tb.active
    tsheet.append(["WMID", "Num", "T", "Tn", "Tx", "U", "P", "FF", "F10", "RRR"])  
    for st in STATIONS:
        fileXLS = st["WMID"] + '_' + sDateStr[0:10] + '_' + eDateStr[0:10] + '.xlsx'
        wb = openpyxl.load_workbook(fileXLS, data_only=True)  
        sheet = wb["Sheet"]
        MR = sheet.max_row - 3
        begList = [st["WMID"], sheet.cell(row = MR+3, column = 2).value, sheet.cell(row = MR, column = 2).value, sheet.cell(row = MR+1, column = 2).value, sheet.cell(row = MR+2, column = 2).value]
        for i in range(3, 7):
            begList.append(sheet.cell(row = MR, column = i).value)   
        begList.append(sheet.cell(row = MR, column = 8).value)    
        tsheet.append(begList)
    tb.save("crimea" + "_" + eDateStr[0:7] + '.xlsx')

writeSummaryXLS()

#for tstamp, observ in weatherData.items():
#    print(tstamp, end = "|")
#    for units in observ:
#        print(str(units['Value']), end = "|")
#    print("\n")

#print (prevObserv(date + "T" + time))      